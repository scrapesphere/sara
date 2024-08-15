import os
import requests
import hashlib
from lxml import etree # type: ignore
import json
import sys
import pdb
import yaml
import importlib.util
from openpyxl import load_workbook # type: ignore
from datetime import datetime
import time

class CommonModule:
    @staticmethod
    def print_error_message(status, info):
        status_message = {
            "status": status,
            "info": info
        }
        json_message = json.dumps(status_message, indent=4)
        print(json_message)

    @staticmethod
    def print_info_message(status, info=None, url=None):
        status_message = {
            "status": status,
            "info": info
        }
        if url is not None:
            status_message["url"] = url
            
        json_message = json.dumps(status_message, indent=4)
        print(json_message)

    @staticmethod
    def get_page_content_hash(url, extended_header=None):
        if url:
            try:
                if extended_header:
                    response = requests.get(url, headers=extended_header, verify=False)
                else:
                    response = requests.get(url, verify=False)

                if response.status_code == 200:
                    result = {
                        "page_doc": response.text,
                        "status_code": response.status_code,
                        "url": url
                    }
                    output_dir = "C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline/cache/"
                    os.makedirs(output_dir, exist_ok=True)  # Ensure the cache directory exists

                    output_file = os.path.join(output_dir, CommonModule.encode(url) + '.html')
                    with open(output_file, 'wb') as file:
                        file.write(response.content)
                    CommonModule.print_info_message("success", output_file, "Page fetched successfully.")
                    return result
                else:
                    #CommonModule.print_error_message("error", url, f"Page fetch failed with status code {response.status_code}.")
                    return {
                        "page_doc": "",
                        "status_code": response.status_code,
                        "url": url,
                        "file_path": output_file
                    }

            except requests.RequestException as e:
                #CommonModule.print_error_message("error", url, f"Page fetch failed: {e}")
                return {
                    "page_doc": "",
                    "status_code": None,
                    "url": url
                }
        else:
            #CommonModule.print_error_message("error", url, "No URL found")
            return {"page_doc": "", "status_code": None, "url": url}

    @staticmethod
    def get_parsed_tree(page_doc):
        try:
            parser = etree.HTMLParser()
            tree = etree.fromstring(page_doc["page_doc"], parser)
            CommonModule.print_info_message("success","Page document parsed successfully.")
            return tree
        except etree.XMLSyntaxError as e:
            CommonModule.print_error_message("error", f"XML Syntax Error: {e}")
            return None
        except Exception as e:
            CommonModule.print_error_message("error", f"Unexpected error: {e}")
            return None
        
    @staticmethod
    def encode(array):
        combined_str = ''.join(array)
        unique_id = hashlib.md5(combined_str.encode()).hexdigest()
        return unique_id
    
    @staticmethod
    def get_value_from_xpath(parsed_tree, xpath_expr, count):
        try:
            elements = parsed_tree.xpath(xpath_expr)
            text_content = [element for element in elements if element]
            #CommonModule.print_error_message("success", "XPath evaluation", f"Text extracted successfully using XPath: {xpath_expr}")
            if count == "all":
                return text_content
            elif count == "first":
                return text_content[0] if text_content else None
        except etree.XPathError as e:
            #CommonModule.print_error_message("error", "XPath evaluation", f"XPath Error: {e}")
            return f"XPath Error: {e}"
        except Exception as e:
            #CommonModule.print_error_message("error", "XPath evaluation", f"Unexpected error: {e}")
            return f"Unexpected error: {e}"

class UrlCollector(CommonModule):
    def __init__(self, base_dir, project_name, site_name):
        self.base_dir = base_dir
        self.output_dir = ""
        self.collector_dir = ""
        self.project_name = project_name
        self.site_name = site_name
        self.count = 0

    @staticmethod
    def write_url_in_txt(self, result_url):
        filepath = os.path.join(self.output_dir, f"{self.site_name}_{self.project_name}.txt")
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, 'a') as file:
            for item in result_url:
                file.write(f"{item}\n")
        CommonModule.print_info_message("success", "Successfully written URLs")

    @staticmethod
    def encode(array):
        combined_str = ''.join(array)
        unique_id = hashlib.md5(combined_str.encode()).hexdigest()
        return unique_id

    @staticmethod
    def enter_count_in_sheet(self):
        if self.count <= 0:
            CommonModule.print_info_message("Failure","No urls to enter in the sheet")
            return
        excel_file = os.path.join(self.base_dir, "url_collector", "url_collector_count_sheet.xlsx")
        sheet_name = "collector_count"
        book = load_workbook(excel_file)
        sheet = book[sheet_name]

        row_num = 2
        while sheet.cell(row=row_num, column=1).value is not None:
            row_num += 1

        array = [time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()), self.site_name, self.project_name]
        unique_id = self.encode(array)
        total_url_count = self.count

        sheet.cell(row=row_num, column=1, value=unique_id)
        sheet.cell(row=row_num, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row=row_num, column=3, value=self.project_name)
        sheet.cell(row=row_num, column=4, value=self.site_name)
        sheet.cell(row=row_num, column=5, value=total_url_count)

        book.save(excel_file)
        CommonModule.print_info_message("success", f"Data written to {excel_file}")

    @staticmethod
    def get_final_url(self, url, depth, current_depth_level, max_depth, module_instance):
        current_depth = depth[f"depth{current_depth_level}"]
        method_name = current_depth["method_name"]
        method_to_call = getattr(module_instance, method_name)
        if method_to_call is None:
            return

        for i in url:
            try:
                result_url = method_to_call(i, depth, current_depth_level)
            except Exception as e:
                continue
            if current_depth_level == max_depth:
                self.write_url_in_txt(self, result_url)
                self.count += len(result_url)
            else:
                self.get_final_url(self, result_url, depth, current_depth_level + 1, max_depth, module_instance)

    @staticmethod
    def main_execution(self):
        try:
            yaml_file_path = os.path.join(self.collector_dir, f"{self.site_name}_{self.project_name}.yml")
            CommonModule.print_info_message("info", f"Loading configuration file: {yaml_file_path}")
            with open(yaml_file_path, 'r') as file:
                depth = yaml.safe_load(file)

            module_path = os.path.join(self.collector_dir, f"{self.site_name}_{self.project_name}.py")

            class_name_in_site_script = f"{self.site_name}_{self.project_name}"
            class_name_in_site_script = ''.join([word.capitalize() for word in class_name_in_site_script.split('_')])
            try:
                spec = importlib.util.spec_from_file_location(class_name_in_site_script, module_path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                SiteClass = getattr(module, class_name_in_site_script)
                site_instance = SiteClass()
            except Exception as e:
                CommonModule.print_error_message("error", f"Error importing module from {module_path}: {e}")
                return

            seed_url = depth["depth0"]["seed_url"]
            if not isinstance(seed_url, list):
                seed_url = [seed_url]

            self.get_final_url(self, seed_url, depth, 0, len(depth) - 1, site_instance)  # Pass the instance
            self.enter_count_in_sheet(self)

        except Exception as e:
            CommonModule.print_error_message("error", f"Unhandled error during execution: {e}")
            raise

    @staticmethod
    def main(self):
        CommonModule.print_info_message("info", f"Starting script execution of url_collector for {site_name}_{project_name}")
        self.output_dir = os.path.join(self.base_dir, f"scrape_output/collector_output/{self.project_name}")
        self.collector_dir = os.path.join(self.base_dir, f"url_collector/{self.project_name}")
        filepath = os.path.join(self.output_dir, f"{self.site_name}_{self.project_name}.txt")
        with open(filepath, 'w') as file:
            file.write('')
        self.main_execution(self)

class UrlFetcher(CommonModule):
    def __init__(self, base_dir, project_name, site_name):
        self.base_dir = base_dir
        self.output_dir = ""
        self.output_dir = ""
        self.project_name = project_name
        self.site_name = site_name

    def encode(self, code):
        unique_id = hashlib.md5(code.encode()).hexdigest()
        return unique_id
    
    def fetch_collector_output(self, project_name, site_name):
        urls = []
        try:
            self.output_dir = os.path.join(self.base_dir, f"scrape_output/collector_output/{project_name}")
            filepath = os.path.normpath(os.path.join(self.output_dir, f"{site_name}_{project_name}.txt"))
            filepath = filepath.replace("//", "/")
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"File not found: {filepath}")
            with open(filepath, 'r') as file:
                urls = [line.strip() for line in file.readlines()]
            args_hash = {}
            yaml_file = os.path.join(self.base_dir, f"url_fetcher/{project_name}/{site_name}_{project_name}.yml")
            with open(yaml_file, 'r') as file:
                args_hash = yaml.safe_load(file)
            module_path = os.path.join(self.base_dir, f"url_fetcher/{project_name}/{site_name}_{project_name}.py")
            class_name_in_site_script = f"{self.site_name}_{self.project_name}"
            class_name_in_site_script = ''.join([word.capitalize() for word in class_name_in_site_script.split('_')])
            try:
                spec = importlib.util.spec_from_file_location(class_name_in_site_script, module_path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                SiteClass = getattr(module, class_name_in_site_script)
                site_instance = SiteClass()
            except Exception as e:
                CommonModule.print_error_message("error", f"Error importing module from {module_path}: {e}")
                return

            for url in urls:
                output_file_path = os.path.normpath(f"C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline/scrape_output/fetcher_output/{project_name}/{site_name}_{project_name}/{UrlFetcher.encode(self,url)}")
                page_content = site_instance.get_page_content(url, args_hash)
                with open(output_file_path, 'w') as file:
                    file.write(str(page_content))
        except FileNotFoundError as e:
            status_message = {
                "status": "Error",
                "file_name": f"{site_name}_{project_name}.txt",
                "project": project_name,
                "site_name": site_name,
                "info": str(e)
            }
            #UrlFetcher.print_status(status_message)
            return []

        except Exception as e:
            status_message = {
                "status": "Error",
                "file_name": f"{site_name}_{project_name}.txt",
                "project": project_name,
                "site_name": site_name,
                "info": f"An unexpected error occurred: {e}"
            }
            print(status_message)
            return []        

if __name__ == "__main__":
    if len(sys.argv) != 4:
        pdb.set_trace()
        print("Usage: python url_fetcher.py <project_name> <site_name>")
        sys.exit(1)
    method_to_execute = sys.argv[1]
    project_name = sys.argv[2]
    site_name = sys.argv[3]
    base_dir = "C:/Users/shanj/OneDrive/Desktop/web-scrapping-pipeline"
    if method_to_execute == "url_collector":
        url_collector = UrlCollector(base_dir,project_name,site_name)
        url_collector.main(url_collector)
    elif method_to_execute == "url_fetcher":
        url_fetcher = UrlFetcher(base_dir,project_name,site_name)
        url_fetcher.fetch_collector_output(project_name, site_name)
    elif method_to_execute == "url_extractor":
        #url_extractor = UrlExtractor(base_dir,project_name,site_name)
        #url_extractor.main(url_extractor)
        print()